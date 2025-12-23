#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script d'analyse détaillée d'un fichier Excel
Analyse lignes 1-54 et colonnes A-BA avec références Excel exactes
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path
import json
from datetime import datetime

def convertir_colonne_excel(num):
    """Convertit un numéro de colonne en lettre Excel (1->A, 27->AA, 53->BA)"""
    return get_column_letter(num)

def analyser_cellule_excel(nom_fichier, sheet_name, ligne, colonne_lettre):
    """Lit une cellule spécifique avec référence Excel"""
    wb = openpyxl.load_workbook(nom_fichier, data_only=True)
    ws = wb[sheet_name]
    return ws[f"{colonne_lettre}{ligne}"].value

def analyser_fichier_excel_detaille(nom_fichier):
    """
    Analyse complète d'un fichier Excel avec références Excel exactes
    Lignes: 1-54, Colonnes: A-BA (53 colonnes)
    """
    print("=" * 100)
    print(f"ANALYSE DÉTAILLÉE DU FICHIER: {nom_fichier}")
    print("=" * 100)
    print(f"Date d'analyse: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # Ouvrir avec openpyxl pour accès direct aux cellules
    try:
        wb = openpyxl.load_workbook(nom_fichier, data_only=True)
    except Exception as e:
        print(f"ERREUR: Impossible d'ouvrir le fichier: {e}")
        return None
    
    # Ouvrir avec pandas pour analyse de données
    xl_file = pd.ExcelFile(nom_fichier)
    
    print(f"Nombre de feuilles: {len(wb.sheetnames)}")
    print(f"Feuilles disponibles: {wb.sheetnames}\n")
    
    resultats = {}
    
    # Analyser chaque feuille
    for sheet_name in wb.sheetnames:
        print("\n" + "=" * 100)
        print(f"FEUILLE: {sheet_name}")
        print("=" * 100)
        
        ws = wb[sheet_name]
        df = pd.read_excel(xl_file, sheet_name=sheet_name, header=None)
        
        # Dimensions Excel réelles
        print(f"\nDimensions Excel: Lignes 1-54, Colonnes A-BA (53 colonnes)")
        print(f"Dimensions pandas: {df.shape[0]} lignes × {df.shape[1]} colonnes")
        
        # Analyser chaque ligne avec référence Excel
        print(f"\n{'='*100}")
        print("ANALYSE LIGNE PAR LIGNE (Références Excel)")
        print(f"{'='*100}\n")
        
        # Analyser les lignes 1-54
        for ligne_excel in range(1, 55):  # Lignes 1 à 54
            ligne_pandas = ligne_excel - 1  # pandas indexe à partir de 0
            
            # Collecter les données non nulles de cette ligne
            donnees_ligne = {}
            cellules_non_vides = []
            
            for col_num in range(1, 54):  # Colonnes A (1) à BA (53)
                col_lettre = convertir_colonne_excel(col_num)
                valeur = ws[f"{col_lettre}{ligne_excel}"].value
                
                if valeur is not None:
                    donnees_ligne[col_lettre] = valeur
                    cellules_non_vides.append(f"{col_lettre}{ligne_excel}: {valeur}")
            
            # Afficher seulement les lignes avec des données
            if cellules_non_vides:
                print(f"\n--- LIGNE {ligne_excel} (Excel) / {ligne_pandas} (pandas) ---")
                print(f"Nombre de cellules non vides: {len(cellules_non_vides)}")
                
                # Afficher toutes les cellules non vides
                for cellule_info in cellules_non_vides[:20]:  # Limiter à 20 pour lisibilité
                    print(f"  {cellule_info}")
                if len(cellules_non_vides) > 20:
                    print(f"  ... et {len(cellules_non_vides) - 20} autres cellules")
                
                # Stocker dans les résultats
                if sheet_name not in resultats:
                    resultats[sheet_name] = {}
                if 'lignes' not in resultats[sheet_name]:
                    resultats[sheet_name]['lignes'] = {}
                resultats[sheet_name]['lignes'][ligne_excel] = {
                    'cellules_non_vides': len(cellules_non_vides),
                    'donnees': donnees_ligne
                }
        
        # Analyser chaque colonne avec référence Excel
        print(f"\n{'='*100}")
        print("ANALYSE COLONNE PAR COLONNE (Références Excel)")
        print(f"{'='*100}\n")
        
        for col_num in range(1, 54):  # Colonnes A à BA
            col_lettre = convertir_colonne_excel(col_num)
            col_pandas = col_num - 1  # pandas indexe à partir de 0
            
            # Analyser la colonne
            valeurs_colonne = []
            for ligne_excel in range(1, 55):
                valeur = ws[f"{col_lettre}{ligne_excel}"].value
                if valeur is not None:
                    valeurs_colonne.append((ligne_excel, valeur))
            
            if valeurs_colonne:
                print(f"\n--- COLONNE {col_lettre} (Excel) / Colonne {col_pandas} (pandas) ---")
                print(f"Nombre de cellules non vides: {len(valeurs_colonne)}")
                print(f"Pourcentage rempli: {len(valeurs_colonne)/54*100:.1f}%")
                
                # Afficher les premières valeurs
                print("Premières valeurs:")
                for ligne_excel, valeur in valeurs_colonne[:10]:
                    print(f"  Ligne {ligne_excel}: {valeur}")
                if len(valeurs_colonne) > 10:
                    print(f"  ... et {len(valeurs_colonne) - 10} autres valeurs")
                
                # Types de données dans cette colonne
                types_valeurs = {}
                for _, valeur in valeurs_colonne:
                    type_val = type(valeur).__name__
                    types_valeurs[type_val] = types_valeurs.get(type_val, 0) + 1
                print(f"Types de données: {types_valeurs}")
                
                # Statistiques si numérique
                valeurs_numeriques = [v for _, v in valeurs_colonne if isinstance(v, (int, float))]
                if valeurs_numeriques:
                    print(f"Statistiques numériques:")
                    print(f"  Min: {min(valeurs_numeriques)}")
                    print(f"  Max: {max(valeurs_numeriques)}")
                    print(f"  Moyenne: {sum(valeurs_numeriques)/len(valeurs_numeriques):.2f}")
                
                # Stocker dans les résultats
                if 'colonnes' not in resultats[sheet_name]:
                    resultats[sheet_name]['colonnes'] = {}
                resultats[sheet_name]['colonnes'][col_lettre] = {
                    'cellules_non_vides': len(valeurs_colonne),
                    'pourcentage_rempli': len(valeurs_colonne)/54*100,
                    'types_donnees': types_valeurs,
                    'valeurs': [(ligne, str(v)) for ligne, v in valeurs_colonne[:50]]  # Limiter à 50
                }
                if valeurs_numeriques:
                    resultats[sheet_name]['colonnes'][col_lettre]['statistiques'] = {
                        'min': min(valeurs_numeriques),
                        'max': max(valeurs_numeriques),
                        'moyenne': sum(valeurs_numeriques)/len(valeurs_numeriques),
                        'count': len(valeurs_numeriques)
                    }
        
        # Analyse structurelle: identifier les sections
        print(f"\n{'='*100}")
        print("ANALYSE STRUCTURELLE - IDENTIFICATION DES SECTIONS")
        print(f"{'='*100}\n")
        
        # Chercher les en-têtes et sections
        sections = {}
        for ligne_excel in range(1, 55):
            # Chercher les cellules avec du texte qui pourrait être un en-tête
            en_tetes_ligne = []
            for col_num in range(1, 54):
                col_lettre = convertir_colonne_excel(col_num)
                valeur = ws[f"{col_lettre}{ligne_excel}"].value
                if isinstance(valeur, str) and valeur.strip():
                    # Vérifier si c'est un en-tête potentiel (texte en majuscules ou mots-clés)
                    if any(mot in valeur.lower() for mot in ['total', 'prix', 'nbr', 'date', 'jour', 'ref', 'client', 'pax', 'adulte', 'enfant']):
                        en_tetes_ligne.append((col_lettre, valeur))
            
            if en_tetes_ligne:
                sections[ligne_excel] = en_tetes_ligne
                print(f"Ligne {ligne_excel} - En-têtes potentiels trouvés:")
                for col, val in en_tetes_ligne:
                    print(f"  {col}{ligne_excel}: {val}")
        
        resultats[sheet_name]['sections'] = sections
        
        # Matrice complète des données
        print(f"\n{'='*100}")
        print("MATRICE COMPLÈTE DES DONNÉES (Toutes les cellules non vides)")
        print(f"{'='*100}\n")
        
        matrice = {}
        total_cellules_non_vides = 0
        for ligne_excel in range(1, 55):
            ligne_data = {}
            for col_num in range(1, 54):
                col_lettre = convertir_colonne_excel(col_num)
                valeur = ws[f"{col_lettre}{ligne_excel}"].value
                if valeur is not None:
                    ligne_data[col_lettre] = valeur
                    total_cellules_non_vides += 1
            if ligne_data:
                matrice[ligne_excel] = ligne_data
        
        print(f"Total de cellules non vides dans la feuille: {total_cellules_non_vides}")
        print(f"Total de cellules possibles: {54 * 53} = {54 * 53}")
        print(f"Pourcentage de remplissage: {total_cellules_non_vides/(54*53)*100:.1f}%")
        
        resultats[sheet_name]['matrice'] = matrice
        resultats[sheet_name]['statistiques_globales'] = {
            'total_cellules_non_vides': total_cellules_non_vides,
            'total_cellules_possibles': 54 * 53,
            'pourcentage_rempli': total_cellules_non_vides/(54*53)*100
        }
    
    # Résumé final
    print("\n" + "=" * 100)
    print("RÉSUMÉ FINAL")
    print("=" * 100)
    print(f"\nFichier analysé: {nom_fichier}")
    print(f"Nombre total de feuilles: {len(wb.sheetnames)}")
    for sheet_name, stats in resultats.items():
        print(f"\nFeuille '{sheet_name}':")
        if 'statistiques_globales' in stats:
            sg = stats['statistiques_globales']
            print(f"  - Cellules non vides: {sg['total_cellules_non_vides']} / {sg['total_cellules_possibles']}")
            print(f"  - Taux de remplissage: {sg['pourcentage_rempli']:.1f}%")
        if 'lignes' in stats:
            print(f"  - Lignes avec données: {len(stats['lignes'])}")
        if 'colonnes' in stats:
            print(f"  - Colonnes avec données: {len(stats['colonnes'])}")
    
    # Sauvegarder les résultats en JSON
    nom_resultat = Path(nom_fichier).stem + "_analyse_detaille.json"
    with open(nom_resultat, 'w', encoding='utf-8') as f:
        json.dump(resultats, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nRésultats détaillés sauvegardés dans: {nom_resultat}")
    
    wb.close()
    return resultats

if __name__ == "__main__":
    fichier = "Bases de datos internos.xlsx"
    
    if not Path(fichier).exists():
        print(f"ERREUR: Le fichier '{fichier}' n'existe pas!")
        exit(1)
    
    resultats = analyser_fichier_excel_detaille(fichier)
    
    print("\n" + "=" * 100)
    print("ANALYSE TERMINÉE")
    print("=" * 100)

