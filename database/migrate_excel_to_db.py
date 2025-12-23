#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de migration des données Excel vers PostgreSQL
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import psycopg2
from psycopg2.extras import execute_values
from datetime import datetime
import sys
from pathlib import Path

# Configuration de la base de données
DB_CONFIG = {
    'host': 'localhost',
    'database': 'cotisation_madagascar',
    'user': 'postgres',
    'password': '2475',  # À modifier selon votre configuration
    'port': 5432
}

def convertir_colonne_excel(num):
    """Convertit un numéro de colonne en lettre Excel"""
    return get_column_letter(num)

def connect_db():
    """Établit la connexion à la base de données"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        print(f"Erreur de connexion à la base de données: {e}")
        sys.exit(1)

def migrer_feuille_excel(nom_fichier, nom_feuille, conn):
    """Migre une feuille Excel vers la base de données"""
    print(f"\n{'='*80}")
    print(f"Migration de la feuille: {nom_feuille}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(nom_fichier, data_only=True)
    ws = wb[nom_feuille]
    cur = conn.cursor()
    
    # Extraire les informations du client depuis le nom de la feuille
    # Format attendu: "Nom X pax" où X est le nombre de personnes
    parts = nom_feuille.split()
    nom_client = parts[0] if len(parts) > 0 else nom_feuille
    nb_pax = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
    
    # Lire les données de base (lignes 1-18)
    ref_client = ws['A2'].value or f"CLIENT_{nom_client}"
    date_cotation = datetime.now().date()  # Par défaut aujourd'hui
    taux_change = ws['L2'].value if ws['L2'].value else 4420
    
    # Extraire le nombre d'adultes, enfants, bébés
    nb_adultes = 0
    nb_enfants = 0
    nb_bebes = 0
    
    for row in range(2, 8):
        label = ws[f'A{row}'].value
        if label == 'Adulte':
            nb_adultes = nb_pax  # Approximation
        elif label == 'Enfant':
            nb_enfants = 0  # À déterminer depuis les données
        elif label == 'Bébé':
            nb_bebes = 0
    
    # Extraire les totaux
    total_ariary = ws['I3'].value or 0
    marge = ws['I4'].value or 0
    total_euro = ws['I6'].value or 0
    
    # Créer ou récupérer le client
    cur.execute("""
        INSERT INTO clients (reference, nom) 
        VALUES (%s, %s)
        ON CONFLICT (reference) DO UPDATE SET nom = EXCLUDED.nom
        RETURNING id
    """, (ref_client, nom_client))
    client_id = cur.fetchone()[0]
    
    # Créer le devis
    reference_devis = f"DEV-{nom_client}-{datetime.now().strftime('%Y%m%d')}"
    cur.execute("""
        INSERT INTO devis (
            client_id, reference, date_cotation, nombre_personnes,
            nombre_adultes, nombre_enfants, nombre_bebes,
            taux_change, total_ariary, total_euro, marge
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (client_id, reference_devis, date_cotation, nb_pax,
          nb_adultes, nb_enfants, nb_bebes, taux_change,
          total_ariary, total_euro, marge))
    devis_id = cur.fetchone()[0]
    
    print(f"Devis créé: {reference_devis} (ID: {devis_id})")
    
    # Migrer les coûts par catégorie (lignes 2-11)
    categories_map = {
        'Pirogue': ('D2', 'E2'),
        'Bateau': ('D3', 'E3'),
        'Location': ('D4', 'E4'),
        'Guidage': ('D5', 'E5'),
        'Carburant': ('D6', 'E6'),
        'Reserves': ('D7', 'E7'),
        'Parcs': ('D8', 'E8'),
        'hebergements': ('D9', 'E9'),
        'Repas': ('D10', 'E10'),
    }
    
    for cat_nom, (cell_nom, cell_montant) in categories_map.items():
        montant = ws[cell_montant].value
        if montant:
            cur.execute("SELECT id FROM categories_couts WHERE nom = %s", (cat_nom,))
            cat_result = cur.fetchone()
            if cat_result:
                cat_id = cat_result[0]
                montant_euro = montant / taux_change if taux_change else 0
                cur.execute("""
                    INSERT INTO couts_devis (devis_id, categorie_id, montant_ariary, montant_euro)
                    VALUES (%s, %s, %s, %s)
                """, (devis_id, cat_id, montant, montant_euro))
    
    # Migrer les jours de voyage (lignes 21-34)
    jours_data = []
    for row in range(21, 35):
        jour_num = ws[f'A{row}'].value
        date_jour = ws[f'B{row}'].value
        destination = ws[f'C{row}'].value
        
        if jour_num and isinstance(jour_num, (int, float)):
            cur.execute("""
                INSERT INTO jours_voyage (devis_id, numero_jour, date_jour, destination, ordre)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            """, (devis_id, int(jour_num), int(date_jour) if date_jour else None, destination, int(jour_num)))
            jour_id = cur.fetchone()[0]
            
            # Migrer les transferts
            transfert_prix = ws[f'F{row}'].value
            if transfert_prix:
                cur.execute("""
                    INSERT INTO transferts (jour_voyage_id, type_transfert, prix_ariary)
                    VALUES (%s, %s, %s)
                """, (jour_id, 'Transfert', transfert_prix))
            
            pirogue_nbr = ws[f'H{row}'].value
            pirogue_prix = ws[f'I{row}'].value
            if pirogue_nbr or pirogue_prix:
                cur.execute("""
                    INSERT INTO transferts (jour_voyage_id, type_transfert, nombre_personnes, prix_ariary)
                    VALUES (%s, %s, %s, %s)
                """, (jour_id, 'Pirogue', pirogue_nbr or 0, pirogue_prix or 0))
            
            bateau_nbr = ws[f'K{row}'].value
            bateau_prix = ws[f'L{row}'].value
            if bateau_nbr or bateau_prix:
                cur.execute("""
                    INSERT INTO transferts (jour_voyage_id, type_transfert, nombre_personnes, prix_ariary)
                    VALUES (%s, %s, %s, %s)
                """, (jour_id, 'Bateau', bateau_nbr or 0, bateau_prix or 0))
            
            # Migrer les locations
            location_nbr = ws[f'N{row}'].value
            location_prix = ws[f'O{row}'].value
            kilometrage = ws[f'P{row}'].value
            if location_nbr or location_prix:
                cur.execute("""
                    INSERT INTO locations_vehicules (jour_voyage_id, type_location, nombre_vehicules, prix_ariary, kilometrage)
                    VALUES (%s, %s, %s, %s, %s)
                """, (jour_id, 'Location sans carburant', location_nbr or 0, location_prix or 0, kilometrage or 0))
            
            # Migrer les guidages
            guidage_nbr = ws[f'R{row}'].value
            guidage_prix = ws[f'S{row}'].value
            if guidage_nbr or guidage_prix:
                cur.execute("""
                    INSERT INTO guidages (jour_voyage_id, nombre_guides, prix_ariary)
                    VALUES (%s, %s, %s)
                """, (jour_id, guidage_nbr or 0, guidage_prix or 0))
            
            # Migrer les réserves/parcs
            reserve_nom = ws[f'T{row}'].value
            reserve_prix = ws[f'U{row}'].value
            reserve_nbr = ws[f'V{row}'].value
            parc_nom = ws[f'Y{row}'].value
            parc_prix = ws[f'Z{row}'].value
            parc_nbr = ws[f'AA{row}'].value
            
            if reserve_nom or reserve_prix:
                cur.execute("""
                    INSERT INTO reserves_parcs (jour_voyage_id, nom_reserve, nombre_personnes, prix_ariary)
                    VALUES (%s, %s, %s, %s)
                """, (jour_id, reserve_nom, reserve_nbr or 0, reserve_prix or 0))
            
            if parc_nom or parc_prix:
                cur.execute("""
                    INSERT INTO reserves_parcs (jour_voyage_id, nom_parc, nombre_personnes, prix_ariary)
                    VALUES (%s, %s, %s, %s)
                """, (jour_id, parc_nom, parc_nbr or 0, parc_prix or 0))
            
            # Migrer les hébergements
            hotel_nom = ws[f'AH{row}'].value
            hotel_prix = ws[f'AI{row}'].value
            hotel_nbr = ws[f'AJ{row}'].value
            transfert_htl = ws[f'AE{row}'].value
            
            if hotel_nom or hotel_prix:
                cur.execute("""
                    INSERT INTO hebergements (jour_voyage_id, type_chambre, nom_hotel, nombre_chambres, prix_ariary, transfert_htl)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (jour_id, 'Double', hotel_nom, hotel_nbr or 0, hotel_prix or 0, transfert_htl or 0))
            
            # Migrer les repas
            vinette_nbr = ws[f'AQ{row}'].value
            vinette_prix = ws[f'AR{row}'].value
            pd_nbr = ws[f'AT{row}'].value
            pd_prix = ws[f'AU{row}'].value
            dn_nbr = ws[f'AW{row}'].value
            dn_prix = ws[f'AX{row}'].value
            dj_nbr = ws[f'AZ{row}'].value
            dj_prix = ws[f'BA{row}'].value
            
            if vinette_nbr or vinette_prix:
                cur.execute("""
                    INSERT INTO repas (jour_voyage_id, type_repas, nombre_personnes, prix_ariary)
                    VALUES (%s, %s, %s, %s)
                """, (jour_id, 'Vinette', vinette_nbr or 0, vinette_prix or 0))
            
            if pd_nbr or pd_prix:
                cur.execute("""
                    INSERT INTO repas (jour_voyage_id, type_repas, nombre_personnes, prix_ariary)
                    VALUES (%s, %s, %s, %s)
                """, (jour_id, 'PD', pd_nbr or 0, pd_prix or 0))
            
            if dn_nbr or dn_prix:
                cur.execute("""
                    INSERT INTO repas (jour_voyage_id, type_repas, nombre_personnes, prix_ariary)
                    VALUES (%s, %s, %s, %s)
                """, (jour_id, 'DN', dn_nbr or 0, dn_prix or 0))
            
            if dj_nbr or dj_prix:
                cur.execute("""
                    INSERT INTO repas (jour_voyage_id, type_repas, nombre_personnes, prix_ariary)
                    VALUES (%s, %s, %s, %s)
                """, (jour_id, 'DJ', dj_nbr or 0, dj_prix or 0))
    
    # Migrer les imprévus (ligne 14)
    imprevu_montant = ws['D14'].value
    imprevu_nbr = ws['E14'].value
    imprevu_prix = ws['F14'].value
    if imprevu_montant or imprevu_prix:
        cur.execute("""
            INSERT INTO imprevus (devis_id, description, nombre, prix_ariary)
            VALUES (%s, %s, %s, %s)
        """, (devis_id, 'Imprevu', imprevu_nbr or 0, imprevu_prix or 0))
    
    conn.commit()
    print(f"✓ Migration terminée pour {nom_feuille}")
    return devis_id

def main():
    """Fonction principale"""
    fichier_excel = "Bases de datos internos.xlsx"
    
    if not Path(fichier_excel).exists():
        print(f"ERREUR: Le fichier '{fichier_excel}' n'existe pas!")
        sys.exit(1)
    
    print("="*80)
    print("MIGRATION DES DONNÉES EXCEL VERS POSTGRESQL")
    print("="*80)
    
    # Connexion à la base de données
    conn = connect_db()
    
    try:
        # Créer les tables si elles n'existent pas
        with open('database/schema.sql', 'r', encoding='utf-8') as f:
            schema_sql = f.read()
            cur = conn.cursor()
            cur.execute(schema_sql)
            conn.commit()
            print("✓ Schéma de base de données créé/vérifié")
        
        # Lire le fichier Excel
        xl_file = pd.ExcelFile(fichier_excel)
        
        # Migrer chaque feuille
        for sheet_name in xl_file.sheet_names:
            try:
                migrer_feuille_excel(fichier_excel, sheet_name, conn)
            except Exception as e:
                print(f"ERREUR lors de la migration de {sheet_name}: {e}")
                conn.rollback()
        
        print("\n" + "="*80)
        print("MIGRATION TERMINÉE AVEC SUCCÈS")
        print("="*80)
        
    except Exception as e:
        print(f"ERREUR: {e}")
        conn.rollback()
    finally:
        conn.close()

if __name__ == "__main__":
    main()

